# This file is part of the pdfsp project
# Copyright (C) 2025 Sermet Pekin
#
# This source code is free software; you can redistribute it and/or
# modify it under the terms of the European Union Public License
# (EUPL), Version 1.2, as published by the European Commission.
#
# You should have received a copy of the EUPL version 1.2 along with this
# program. If not, you can obtain it at:
# <https://joinup.ec.europa.eu/collection/eupl/eupl-text-eupl-12>.
#
# This source code is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the
# European Union Public License for more details.
#
# Alternatively, if agreed upon, you may use this code under any later
# version of the EUPL published by the European Commission.
import pdfplumber
import pandas as pd
from dataclasses import dataclass
import traceback
import os
from pathlib import Path


@dataclass
class DataFrame:
    df: pd.DataFrame
    path: Path
    out: str = None
    page: int = 1
    index: int = 1
    extra: tuple = ()
    name: str = ""

    def __post_init__(self):
        if self.out is None:
            self.out = "Output"
        self.out = Path(self.out)
        self.df = self.make_unique_cols(self.df)
        self.name = Path(self.path).stem.split(".pdf")[0]

    def make_unique_cols(self, df: pd.DataFrame) -> pd.DataFrame:
        cols = [str(x) for x in df.columns]
        df.columns = self.make_unique(cols)
        return df

    def make_unique(self, cols: list[str]) -> list[str]:
        a = []
        b = []
        for i, col in enumerate(cols):
            ucol = col
            if col in a:
                col = col + str(i)
                ucol = f"{col}-{i}"
            a.append(col)
            b.append(ucol)
        return b

    def get_file_name(self) -> str:
        return f"[{self.name}]-Page {self.page}-T {self.index}.xlsx"

    def create_dir(self) -> None:
        os.makedirs(self.out, exist_ok=True)

    def write(self):
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows

        self.create_dir()
        file_name = self.get_file_name()
        wb = Workbook()
        ws = wb.active
        title = f"{self.name}-Table-{self.index}"
        ws.title = title
        for r_idx, row in enumerate(
            dataframe_to_rows(self.df, index=False, header=True), start=1
        ):
            ws.append(row)
        footnote_row = len(self.df) + 4
        ws.cell(row=footnote_row, column=1, value=f"Footnote: {title} ")
        paragraph_row = footnote_row + 4
        ws.cell(
            row=paragraph_row,
            column=1,
            value=f"This table was extracted from {self.path} with pdfsp package.",
        )
        wb.save(self.out / file_name)
        print(f"[writing table] {file_name}")


def check_folder(folder: Path) -> bool:
    """Check if the folder exists and is a directory."""
    folder = Path(folder)
    if not folder.exists():
        print(f"Folder `{folder}` does not exist.")
        return False
    if not folder.is_dir():
        print(f"`{folder}` is not a directory.")
        return False
    return True


def get_pdf_files(folder: Path = None, out: str = None) -> list[str]:
    """Get all PDF files in the specified folder."""
    if folder is None:
        folder = Path(".")
    if out is None:
        out = Path(".")

    if not check_folder(folder):
        return []

    print(f"Searching for PDF files in `{folder}`")

    files = [Path(folder) / x for x in os.listdir(folder) if x.endswith(".pdf")]
    if not files:
        print(f"No PDF files found in `{folder}`")
        return []
    print(f"Found {len(files)} PDF files in `{folder}`")
    return files


def extract_tables_from_pdf(pdf_path, out: Path = None) -> DataFrame:
    """Extract tables from a PDF file."""

    with pdfplumber.open(pdf_path) as pdf:
        print(f"""Extracting tables from `{pdf_path}`""")
        for i, page in enumerate(pdf.pages, start=1):

            tables = page.extract_tables()
            for index, table in enumerate(tables):
                df = pd.DataFrame(table[1:], columns=table[0])
                yield DataFrame(df, pdf_path, out, page=i, index=index + 1)


def write_dfs(pdf_files: list[Path], out: Path = None):
    """Write DataFrames to Excel files."""
    for pdf_file in pdf_files:
        for df in extract_tables_from_pdf(pdf_file, out):
            print(f"Writing table from `{df.path}`")
            df.write()


def extract_tables(folder: Path = None, out: str = None):
    """Extract tables from all PDF files in the specified folder."""
    for file in get_pdf_files(folder, out):
        pdf_ = extract_tables_from_pdf(file, out)
        for df in pdf_:
            df.write()

    files = get_pdf_files(folder, out)
    write_dfs(files, out)
