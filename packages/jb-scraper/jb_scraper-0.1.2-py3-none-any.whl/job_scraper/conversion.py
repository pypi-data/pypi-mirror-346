from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import load_workbook
import os
import inspect


# This function will get the path of
# the module executing the script.
def get_root_caller_path():

    stack = inspect.stack()
    root_caller = stack[-1]
    root_file = root_caller.filename

    return root_file


def list_conversion(job_list: list, archive_name: str, sheet_name: str):

    ROOT_FOLDER = Path(get_root_caller_path()).parent
    WORKBOOK_PATH = ROOT_FOLDER / (archive_name + ".xlsx")

    if os.path.exists(WORKBOOK_PATH):
        workbook = load_workbook(WORKBOOK_PATH)

        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]  # type: ignore

        workbook.create_sheet(sheet_name, len(workbook.sheetnames))
        worksheet: Worksheet = workbook[sheet_name]  # type: ignore

        worksheet.cell(1, 1, "Job Name")
        worksheet.cell(1, 2, "Job Link")

        for job in job_list:
            if job[0] in worksheet:
                pass
            else:
                worksheet.append(job)

        workbook.save(WORKBOOK_PATH)

    else:
        workbook = Workbook()
        del workbook["Sheet"]  # type: ignore

        workbook.create_sheet(sheet_name, len(workbook.sheetnames))

        worksheet: Worksheet = workbook[sheet_name]  # type: ignore

        worksheet.cell(1, 1, "Job Name")
        worksheet.cell(1, 2, "Job Link")

        for job in job_list:
            if job[0] in worksheet:
                pass
            else:
                worksheet.append(job)

        workbook.save(WORKBOOK_PATH)
