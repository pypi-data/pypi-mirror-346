
import calendar
import datetime
from enum import Enum, auto
import os
import sqlite3
import numpy as np
from openpyxl import load_workbook
import pandas as pd
import logging
#from database.orm_management import Motor, motor_to_acmachine, motor_to_df
from .electrical_machines import ACMachine,DCMachine
from .sqlite_tools import QuaryOutFormat, SqliteTools

logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

    
class DataScope(Enum):
    '''display options'''
    consumption_L1 =auto()
    consumption_L2 =  auto()
    consumption_plant = auto()
    consumption_area =auto()
    consumption_area_no_bg = auto()  
    
class PowerAna:
    '''
    Power analysis class based on monthly power readings and daily production report
    '''
    def __init__(self,power_unit_cost=1.015,power_max_use_cost=50) -> None:
        self.power_unit_cost = power_unit_cost
        self.power_max_use_cost = power_max_use_cost
        logger.info('loading power consumption data...')
        self.__load_consumption_data()
        logger.info('loading production reports data...')
        self.__load_production_data()
        logger.info('loading motors data...')
        self.__load_motors_db(as_df=True)
        logger.info('done...')
        
    def __locate_tables(self,search_term:str,coordinate_criteria:str,sheet_name:str='Sheet1') -> list:
        '''
        locate each table beginning
        
        Parameters
        ----------
        search_term : str .
            string to search for. 
        coordinate_criteria : str .
            coordinate criteria
        sheet_name : str (default: 'Sheet1'). 
            spread sheet file name (should be in current direcotry)
        Returns:
        -------
            none
        '''
        table_locations = []
        for row in self.wb[sheet_name].iter_rows():
            for cell in row:
                if (cell.value == search_term and cell.coordinate.startswith(coordinate_criteria)) :
                    table_locations.append(cell.coordinate)
        return table_locations
    
    def __load_consumption_data(self,wb_name:str='SS2_2024.xlsx'):
        '''
        load power data table from monthly power readings
        '''
        db_path = "c:\\Users\\walie\\.vscode\\projects\\python\\titan\\power_analysis\\"
        self.wb = load_workbook(filename= f'{db_path}{wb_name}',read_only=True)
        tables_locs = [s[1:] for s in self.__locate_tables(sheet_name='ورقة1',search_term='F11',coordinate_criteria='A')]
        tables_locs = [int(s) for s in tables_locs]
        columns_names = ['feeder_no','feeder_no_1' , 'feeder' , 'description' , 'new_read' , 'old_read' , 'diff' ,'base_cons', 'factor' , 'loss_factor' , 'consumptions']
        self.consumption = [pd.read_excel(f'{db_path}SS2_2024.xlsx',sheet_name='ورقة1' , skiprows= p-2 , nrows= 23, usecols="A:K" , names=columns_names)
                        for p in tables_locs]
        
        for x in range(0,len(self.consumption)-1,1):
            self.consumption[x]['month'] = x+1
            
        #self.consumption = [x.drop(['feeder_no_1'],axis=1) for x in self.consumption]
        
        #self.consumption = pd.concat(self.consumption)
        [x.fillna({'feeder': 'coal2'},inplace=True) for x in self.consumption] 
        [x.drop(columns=['feeder_no','feeder_no_1','description','new_read','old_read','diff','base_cons','factor','loss_factor'],inplace=True ) for x in self.consumption]
        
        columns = self.consumption[0]['feeder'].values
        self.consumption = [x['consumptions'].to_frame().T for x in self.consumption]
        self.consumption = pd.concat(self.consumption)
        self.consumption.columns = columns        
        self.consumption.index = [datetime.date(2024,month_num,1).strftime("%b") for month_num in range(1,len(tables_locs)+1)]
        cols_list = self.consumption.columns.to_list()
        index_1 = cols_list.index('ER-6',1)
        cols_list[index_1] = 'ER-6_1'
        index_1 = cols_list.index('ER-4',1)
        cols_list[index_1] = 'ER-4_1'
        self.consumption.columns = cols_list
        #self.consumption.reset_index(inplace=True)
        self.consumption_L2 = self.consumption.filter(items=['coal2',	'ER-321"',	'ER900TF05',	'ER-230',	'ER-430',	'ER-320',	'ER-321'])

        self.consumption_L1 = self.consumption.filter(items=['DT031',	'ER-2',	'ER-6_1',	'ER-9',	'ER-8',	'ER-4_1',	'ER-5',	'ER-7',	'ER-6',	'ER-4',	'ER-1',	'DT032','SS2-DH10',	'SS2-DH15',	'DT001',	'DT002'])
        
        new_locs = [x+22+6 for x in tables_locs]
        new_indices = ['peak_max_mw' , 'month_max_mw' , 'kwh' , 'kvar' , 'kwh_peak' , 'kvar_peak']
        
        self.consumption_total = [pd.read_excel(f'{db_path}SS2_2024.xlsx',sheet_name='ورقة1' , skiprows= p-1 , nrows= 6, usecols="D:E",names=['desc','value'])
                        for p in new_locs]
        
        p_tot_list = []
        #x_count = 1
        for x in self.consumption_total :
            newitem = pd.Series(data=x['value'].values,dtype=float,index=new_indices)
            #newitem['month'] = x_count
            p_tot_list.append(newitem)
            #x_count = x_count+1
        #self.consumption_total  = [pd.Series(data=x['value'].values,dtype=float,index=new_indices) for x in self.consumption_total]
        self.consumption_total = pd.concat(p_tot_list,axis=1)
        
        month_dic = {month_num: datetime.date(2024,month_num,1).strftime("%B") for month_num in range(1,len(self.consumption_total.keys())+1)}
        month_reorder_dic = {month_num: month_num+1 for month_num in range(0,len(self.consumption_total.keys()))}
        self.consumption_total.rename(columns=month_reorder_dic,inplace=True)
        self.consumption_total.rename(columns=month_dic,inplace=True)
        self.consumption_total = self.consumption_total.T
        self.consumption_total['pf']  = round(np.cos(np.arctan(self.consumption_total['kvar'] / self.consumption_total['kwh'])),3)
        
        self.monthdays = [calendar.monthrange(2024,i)[1] for i in range(1,self.consumption_total.index.size+1)]
        self.consumption_total['av_kw']  = (self.consumption_total['kwh']  / self.monthdays)/24
        self.consumption_total['av_kvar']  = (self.consumption_total['kvar']  / self.monthdays)/24
        self.consumption_total['load_factor']  = self.consumption_total['av_kw']  / (self.consumption_total['month_max_mw']*1000)
        # Forming Equipement dataframe 
        self.consumption_aquipment  = pd.DataFrame()  
        self.consumption_aquipment['Crusher'] = self.consumption['ER-1']
        self.consumption_aquipment['RMIX'] = self.consumption['ER-2']
        self.consumption_aquipment['RM1'] = self.consumption['ER-4'] + self.consumption['ER-4_1']  
        self.consumption_aquipment['RM2'] = self.consumption['ER-230']
        self.consumption_aquipment['SF1'] = self.consumption['SS2-DH10'] + self.consumption['SS2-DH15']
        self.consumption_aquipment['SF2'] = self.consumption['coal2']
        self.consumption_aquipment['KILN1'] = self.consumption['ER-5'] 
        self.consumption_aquipment['KILN2'] = self.consumption['ER-320'] + self.consumption['ER-321'] 
        self.consumption_aquipment['CMAB'] = self.consumption['ER-6'] + self.consumption['ER-6_1'] 
        self.consumption_aquipment['CMC'] = self.consumption['ER-430'] 
        self.consumption_aquipment['Packing'] = self.consumption['ER-7'] 
        self.consumption_aquipment['Utilities1'] = self.consumption['DT031'] + self.consumption['ER-9']  + self.consumption['ER-8']  + self.consumption['DT032'] +self.consumption['DT001']  + self.consumption['DT002'] 
        self.consumption_aquipment['Utilities2'] = self.consumption['ER900TF05'] + self.consumption['ER321"'] 
        # cost
        self.power_unit_cost_peak = self.power_unit_cost * 1.5
    
    def __load_production_data(self,wb_name:str='Daily Report.xlsx'):
        db_path = "c:\\Users\\walie\\.vscode\\projects\\python\\titan\\power_analysis\\"
        self.wb = load_workbook(filename= f'{db_path}{wb_name}',read_only=True)
        tables_starts = [s[1:] for s in self.__locate_tables(sheet_name='Stoppages reason',search_term='Date',coordinate_criteria='A')]
        tables_starts = [int(s)+3 for s in tables_starts]
        tables_rows_count = [calendar.monthrange(year=2024,month=x)[1] for x in range(1,len(tables_starts)+1)]
        table_count = [(tables_starts[i],tables_rows_count[i]) for i in range(min(len(tables_starts),len(tables_rows_count)))]
        columns_names = ['operation_hrs' , 'breakdown_duration' , 'breakdown_reason' , 'schedule_duration' , 'schedule_reason' , 'total_production']
        used_months = [datetime.datetime(2024,x,1).strftime('%b') for x in range(1,len(table_count)+1)]
        # report_columns_used = [("rma","B:G" , 
        #                         ("rmb","I:N") ,
        #                         ("rmc","P:U") ,
        #                         ("sf1","W:AB") ,
        #                         ("sf2","AD:AI") , 
        #                         ("kiln1","AK:AP") , 
        #                         ("kiln2","AQ:AV") , 
        #                         ("cma","AW:BB") , 
        #                         ("cmb","BD:BI" ), 
        #                         ("cmc","BK:BP") , 
        #                         ("rdf","BT:BY")
        #                         ]
        report_columns_used = [("rma", (2,7)) , 
                                ("rmb",(9,14)) ,
                                ("rmc",(16,21)) ,
                                ("sf1",(23,28)) ,
                                ("sf2",(30,35)) , 
                                ("kiln1",(37,42)) , 
                                ("kiln2",(43,48)) , 
                                ("cma",(49,54)) , 
                                ("cmb",(56,61) ), 
                                ("cmc",(63,68)) , 
                                ("rdf",(72,77))
                                ]

       
        self.reports = {}
        data = []
        months = []
        # loop though areas
        for area_name,area_loc in report_columns_used:
            col_min,col_max=area_loc
            # llop through months
            for p,q in table_count:
                for day in self.wb['Stoppages reason'].iter_rows(min_row= p , max_row= p+q-1 , min_col= col_min, max_col=col_max,values_only=True): 
                    data.append(day)
                month =  pd.DataFrame(data=data, columns=columns_names)
                month.set_index([pd.Index(range(1,len(month)+1))],inplace=True)
                month.fillna({'operation_hrs':0,'breakdown_duration':0,'schedule_duration':0,'total_production':0},inplace=True)
                month.fillna({'breakdown_reason':'','schedule_reason':''},inplace=True)
                data.clear()
                months.append(month)
            self.reports[area_name] = dict(zip(used_months,months))    
            months.clear()
        data.clear()
        months.clear()    
        for p,q in table_count:
            for day in self.wb['Stoppages reason'].iter_rows(min_row= p , max_row= p+q-1 , min_col= 71,max_col=71, values_only=True): 
                data.append(day)
            month =  pd.DataFrame(data=data, columns=['dispatch'],index=[pd.Index(range(1,len(data)+1))])
            month.fillna({'dispatch':0},inplace=True)
            data.clear()
            months.append(month) 
        self.reports['packing'] = dict(zip(used_months,months))    
            
        
        # for area_name,area_loc in report_columns_used:
        #     self.reports[area_name] =[pd.read_excel(f'{db_path}{wb_name}',sheet_name='Stoppages reason' , skiprows= p-3 , nrows= q, usecols=area_loc , names=columns_names) for p,q in table_count]
        
        # self.reports['rma'] =[pd.read_excel(f'{db_path}{wb_name}',sheet_name='Stoppages reason' , skiprows= p-3 , nrows= q, usecols="B:G" , names=columns_names) for p,q in table_count]
        
        # self.reports['rmb'] = [pd.read_excel(f'{db_path}{wb_name}',sheet_name='Stoppages reason' , skiprows= p-3 , nrows= q, usecols="I:N" , names=columns_names) for p,q in table_count]
        # # self.reports['rmb']  = [x.reindex(range(1,len(x)+1)) for x in self.reports['rmb']]         
        # # self.reports['rmb'] = dict( zip(used_months,self.reports['rmb'] ))
        
        # self.reports['rmc'] = [pd.read_excel(f'{db_path}{wb_name}',sheet_name='Stoppages reason' , skiprows= p-3 , nrows= q, usecols="P:U" , names=columns_names) for p,q in table_count]
        
        # self.reports['sf1'] = [pd.read_excel(f'{db_path}{wb_name}',sheet_name='Stoppages reason' , skiprows= p-3 , nrows= q, usecols="W:AB" , names=columns_names) for p,q in table_count]
        
        # self.reports['sf2'] = [pd.read_excel(f'{db_path}{wb_name}',sheet_name='Stoppages reason' , skiprows= p-3 , nrows= q, usecols="AD:AI" , names=columns_names) for p,q in table_count]
        
        # self.reports['kiln1'] = [pd.read_excel(f'{db_path}{wb_name}',sheet_name='Stoppages reason' , skiprows= p-3 , nrows= q, usecols="AK:AP" , names=columns_names) for p,q in table_count]
        
        # self.reports['kiln2'] = [pd.read_excel(f'{db_path}{wb_name}',sheet_name='Stoppages reason' , skiprows= p-3 , nrows= q, usecols="AQ:AV" , names=columns_names) for p,q in table_count]
        
        # self.reports['cma'] = [pd.read_excel(f'{db_path}{wb_name}',sheet_name='Stoppages reason' , skiprows= p-3 , nrows= q, usecols="AW:BB" , names=columns_names) for p,q in table_count]
        
        # self.reports['cmb'] = [pd.read_excel(f'{db_path}{wb_name}',sheet_name='Stoppages reason' , skiprows= p-3 , nrows= q, usecols="BD:BI" , names=columns_names) for p,q in table_count]
        
        # self.reports['cmc'] = [pd.read_excel(f'{db_path}{wb_name}',sheet_name='Stoppages reason' , skiprows= p-3 , nrows= q, usecols="BL:BQ" , names=columns_names) for p,q in table_count]

        # self.reports['rdf'] = [pd.read_excel(f'{db_path}{wb_name}',sheet_name='Stoppages reason' , skiprows= p-3 , nrows= q, usecols="BT:BY" , names=columns_names) for p,q in table_count]

        # for area_key,area_value in self.reports.items():
        #     [ x.set_index([pd.Index(range(1,len(x)+1))],inplace=True) for x in area_value]
        #     self.reports[area_key]  =  dict( zip(used_months,area_value))             
        
        final = 0
        
    def __load_motors_db(self) -> pd.DataFrame|None:
        '''
        Loads motors database 
        '''               
        if db_path is None:
                base_dir = os.path.dirname(os.path.abspath(__file__))
                db_path = os.path.join(base_dir, 'motors.db')
        db = SqliteTools(db_path)
        if db is None:
            raise sqlite3.Error(f"Error connecting to motors database")
        db.connect_db()
        self.motors = db.read_table('motor' , QuaryOutFormat = QuaryOutFormat.df )      
        
    def seach_in_range(self,search_term:str):
        # Define the range
        min_row = 2
        max_row = 10
        min_col = 1
        max_col = 5
        for row in self.wb['ورقة1'].iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                if cell.value == search_term:
                    print(f"Found {search_term} in cell {cell.coordinate}")

    def power_consumption(self,Line = 0):
        '''Total power consumption at peak and normal operation hours'''
        total_power_consumption = self.consumption_total.kwh.sum()
        total_power_consumption_peak = self.consumption_total.kwh_peak.sum()
        l1_p = self.consumption_L1.sum().sum()
        l2_p = self.consumption_L2.sum().sum()
        L1_to_L2 = l1_p/l2_p
        if Line == 1:
            return total_power_consumption* (1-L1_to_L2),total_power_consumption_peak *(1-L1_to_L2)
        elif Line == 2:
            return [total_power_consumption * L1_to_L2,total_power_consumption_peak * L1_to_L2]
        else :
            return [total_power_consumption,total_power_consumption_peak]
    
    def power_cost(self,Line = 0):
        power_consumption,power_consumption_peak = self.power_consumption(Line=Line)
        power_total_cost = self.power_unit_cost*power_consumption + self.power_unit_cost_peak*power_consumption_peak
        return power_total_cost
    
    def power_cost_max_use(self):
        return self.power_max_use_cost * (self.consumption_total['month_max_mw'][0:2].max() 
                                                    + self.consumption_total['month_max_mw'][3:5].max()
                                                    + self.consumption_total['month_max_mw'][6:8].max()
                                                    + self.consumption_total['month_max_mw'][9:11].max())
            
    def load_factor_mean(self,month_start= 1,month_end = 12):
        return self.consumption_total['load_factor'][month_start-1:month_end].mean()
        
    def penalty(self,pf,kwh,kwh_peak , pf_set = 0.92):
        if pf_set >= pf:
            return 0.5 * (pf_set-pf)* self.power_unit_cost * kwh  + 0.5 * (pf_set-pf) * self.power_unit_cost_peak * kwh_peak 
        else:
            return 0
        
    def power_uv_lossess(self,motor_code:str,voltage:int) -> float:
        '''
        Steady-state operation of induction motor at undervoltage
        Parameters
        ----------
        motor_code : str 
            selected motor from motor database
        voltage : int 
            undervoltage value
        Returns:
        -------
            power lossses in kw
        '''
        db = SqliteTools('motors.db')
        if db is None:
            raise sqlite3.Error(f"Error connecting to motors database")
        db.connect_db()
        # TODO: modify error in this line
        ac_motor = db.get_items_by_quary(table_name= 'motors' ,quary_column= 'code',quary_value=motor_code)
        ac_motor  = ACMachine()
        #Based on the small-slip approximation s ∝ 1 /V^2, the new slip at the low voltage is
        return ac_motor.undervoltage_lossess(volt=voltage)

class TaariffEgypt:
    
    def __init__(self):
        '''
            تعريفة بيع الكهرباء للعام المالي 2025/2024 * بدءاً من استهلاك شهر أغسطس
            Source: https://egyptera.org/ar/TarrifAug2024.aspx
        '''
        self.consumption_tariffs_residential = [
            (0, 50, 1),   # Up to 50 kWh
            (51, 100, 2), 
            (101,200, 6),
            (201,350, 11),
            (351,560, 15),
            (651,1000, 25),
            (1001, float('inf'), 40)  # Above 1000 kWh
        ]
        self.consumption_tariffs_commercial_shops = [
            (0, 100, 5), 
            (101,250, 15),
            (251,600, 20),
            (601,1000, 25),
            (1001, float('inf'), 40)  # Above 1000 kWh
        ]
        self.consumption_tariffs_380V = 2.34
        self.consumption_tariffs_22kV = 1.94
        self.consumption_tariffs_66kV = 1.74
    # calculate days in a given month and year
    def __days_in_month(self,year, month):
        if month == 12:
            return 31
        return (datetime(year, month + 1, 1) - datetime(year, month, 1)).days

    def __calculate_cost(self,consumption_kwh):
        ''''Calculate the cost of electrical energy consumption based on the consumption in kWh'''
        cost = 0
        for lower, upper, tariff in self.consumption_tariffs_residential:
            if consumption_kwh > lower:
                usage = min(consumption_kwh, upper) - lower
                cost += usage * tariff
        return cost

    # Function to calculate the electrical energy consumption cost
    def calculate_energy_cost(self,year, daily_consumption_kwh):
        '''Calculate the cost of electrical energy consumption for a given year and daily consumption in kWh'''
        total_cost = 0
        cost_result = []
        for month in range(1, 13):
            days = self.__days_in_month(year, month)
            monthly_consumption = daily_consumption_kwh * days
            cost = self.__calculate_cost(monthly_consumption)
            total_cost += cost
            cost_result.append((month, days, monthly_consumption, cost))
            #print(f"Month: {month}, Days: {days}, Consumption: {monthly_consumption} kWh, Cost: {cost:.2f} LE")
        #print(f"Total annual cost: {total_cost:.2f} LE")
        return total_cost,cost_result

# Example usage
# year = 2024
# daily_consumption_kwh = 100  # Example daily consumption in kWh
# taariff = TaariffEgypt()
# total,monthly = taariff.calculate_energy_cost(year, daily_consumption_kwh)
# print(monthly)
    
# pa = PowerAna()
# print(pa.reports['rma']['Feb'])
# print('Line one')
# print(pa.consumption)
# print(pa.consumption_aquipment['RM1'])
# print('Line two')
# print(pa.consumption_ss2_L2.head(3))
# print('pf')
# print(pa.consumption_total)