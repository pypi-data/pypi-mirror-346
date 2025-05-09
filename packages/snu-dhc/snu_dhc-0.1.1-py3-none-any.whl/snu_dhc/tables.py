#!/usr/bin/env python
# coding: utf-8

# In[ ]:


import numpy as np
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from scipy.stats import ttest_ind, mannwhitneyu, chi2_contingency, f_oneway, kruskal, normaltest, fisher_exact
from typing import Union, Literal
import warnings

class DemoTable:
    def __init__(self, df: pd.DataFrame, *,
                 table_name: str,
                 variables: list,
                 group_variable: str = None,
                 group_labels: dict = None,
                 show_total: bool = True,
                 percent_decimals: int = 1,
                 thousands_sep: bool = True,
                 show_p_values: bool = False,
                 p_value_decimals: Union[int, Literal["auto"]] = "auto"):
        """
        Initializes a DemoTable for generating baseline characteristics Excel tables.

        Parameters:
        -----------
        df : pd.DataFrame
            Input dataframe containing the raw data.

        table_name : str
            Title to be shown at the top of the Excel sheet.

        variables : list of dict
            List of variables to include in the table.
            Each dict must include:
                - "var" : str  → column name in df
                - "name" : str  → printed variable name
                - "type" : "categorical" or "continuous"
                - For continuous:
                    - "stat" : "mean" or "median"
                    - Optional: "decimals" : int (decimal places for mean/std or median/IQR)
                - For categorical:
                    - "class_labels" : dict {value: label_to_display}

            Example:
            variables = [
                {
                    "var": "age", "name": "Age", "type": "continuous", "stat": "mean", "decimals": 1
                },
                {
                    "var": "sex", "name": "Sex", "type": "categorical", "class_labels": {1: "Male", 0: "Female"}
                }
            ]

        group_variable : str, optional
            Column name to use for group comparison (e.g., 'treatment_group').

        group_labels : dict, optional
            Dictionary mapping group values to display names.
            Example: {1: "Treatment", 0: "Control"}

        show_total : bool, default=True
            Whether to include a "Total" column.

        percent_decimals : int, default=1
            Number of decimal places to show in percentages.

        thousands_sep : bool, default=True
            Whether to use commas in large numbers (e.g., 1,234).

        show_p_values : bool, default=False
            Whether to include a column for p-values.

        p_value_decimals : int or "auto", default="auto"
            Number of decimal places for p-values.
            - If int: fixed number of decimals.
            - If "auto": 
                * p < 0.01 → show 3 decimals
                * 0.01 ≤ p < 0.045 → show 2 decimals
                * 0.045 ≤ p < 0.055 → show 3 decimals
                * p ≥ 0.055 → show 2 decimals
                * p < 0.001 → "<0.001"
                * p > 0.99 → ">0.99"
        """
        
        self.df = df
        self.table_name = table_name
        self.variables = variables
        self.group_variable = group_variable
        self.group_labels = group_labels or {}
        self.show_total = show_total
        self.percent_decimals = percent_decimals
        self.thousands_sep = thousands_sep
        self.show_p_values = show_p_values
        self.p_value_decimals = p_value_decimals

        if not isinstance(df, pd.DataFrame):
            raise TypeError("df must be a pandas DataFrame.")
        
        if not isinstance(variables, list) or not all(isinstance(v, dict) for v in variables):
            raise TypeError("variables must be a list of dictionaries.")
        
        if not isinstance(percent_decimals, int):
            raise TypeError("percent_decimals must be an integer.")
        
        if not (isinstance(p_value_decimals, int) or p_value_decimals == "auto"):
            raise ValueError("p_value_decimals must be an integer or 'auto'.")

    def save(self, filepath: str):
        table_df = self._build_table()
        self._write_excel(table_df, filepath)

    def _format_number(self, num, decimals):
        if pd.isnull(num):
            return ""
        fmt = f"{{:,.{decimals}f}}" if self.thousands_sep else f"{{:.{decimals}f}}"
        return fmt.format(num)

    def _categorical_summary(self, series, class_labels):
        counts = series.value_counts(dropna=False)
        total = counts.sum()
        summary = {}
        for cls, label in class_labels.items():
            n = counts.get(cls, 0)
            pct = 100 * n / total if total > 0 else 0
            formatted = f"{n:,} ({pct:.{self.percent_decimals}f}%)" if self.thousands_sep else f"{n} ({pct:.{self.percent_decimals}f}%)"
            summary[label] = formatted
        return summary

    def _continuous_summary(self, series, stat, decimals):
        series = series.dropna()
        if len(series) == 0:
            return ""
        if stat == 'mean':
            mean = np.mean(series)
            std = np.std(series, ddof=1)
            return f"{self._format_number(mean, decimals)} ({self._format_number(std, decimals)})"
        elif stat == 'median':
            med = np.median(series)
            q1 = np.percentile(series, 25)
            q3 = np.percentile(series, 75)
            return f"{self._format_number(med, decimals)} ({self._format_number(q1, decimals)}–{self._format_number(q3, decimals)})"
        else:
            return ""

    def _format_p(self, p):
        if pd.isnull(p):
            return ""
    
        if p < 0.001:
            return "<0.001"
        if p > 0.99:
            return ">0.99"
    
        if isinstance(self.p_value_decimals, int):
            return f"{p:.{self.p_value_decimals}f}"
    
        # auto mode
        if p < 0.01:
            return f"{p:.3f}"
        elif 0.045 <= p < 0.055:
            return f"{p:.3f}"  # avoid ambiguity near 0.05
        else:
            return f"{p:.2f}"

    def _compute_p_value(self, var_cfg):
        if not self.group_variable:
            return None
    
        group_vals = [g for g in self.group_labels if g in self.df[self.group_variable].unique()]
        grouped_data = [self.df[self.df[self.group_variable] == g][var_cfg['var']].dropna() for g in group_vals]
    
        if var_cfg["type"] == "continuous":
            if var_cfg["stat"] == "mean":
                if len(grouped_data) == 2:
                    # Normality check
                    norm1 = normaltest(grouped_data[0])[1] if len(grouped_data[0]) >= 8 else 1.0
                    norm2 = normaltest(grouped_data[1])[1] if len(grouped_data[1]) >= 8 else 1.0
                    if norm1 < 0.05 or norm2 < 0.05:
                        warnings.warn(
                            f"Non-normal distribution detected for variable '{var_cfg['name']}'. Consider using Mann–Whitney U test instead of t-test.",
                            UserWarning
                        )
                    _, p = ttest_ind(grouped_data[0], grouped_data[1], equal_var=False)
                elif len(grouped_data) > 2:
                    _, p = f_oneway(*grouped_data)
                else:
                    p = None
            elif var_cfg["stat"] == "median":
                if len(grouped_data) == 2:
                    _, p = mannwhitneyu(grouped_data[0], grouped_data[1], alternative='two-sided')
                elif len(grouped_data) > 2:
                    _, p = kruskal(*grouped_data)
                else:
                    p = None
            else:
                p = None
    
        elif var_cfg["type"] == "categorical":
            contingency = pd.crosstab(self.df[var_cfg['var']], self.df[self.group_variable])
            if contingency.shape[0] > 1 and contingency.shape[1] > 1:
                if (contingency < 5).any().any():
                    warnings.warn(
                        f"Low expected cell count detected for variable '{var_cfg['name']}'. Consider using Fisher's exact test instead of chi-square test.",
                        UserWarning
                    )
                try:
                    _, p, _, _ = chi2_contingency(contingency)
                except:
                    p = None
            else:
                p = None
        else:
            p = None
        return p

    def _build_table(self):
        rows = []
        groups = []
    
        # Determine groups
        if self.group_variable:
            groups = [g for g in self.group_labels if g in self.df[self.group_variable].unique()]
            if self.show_total:
                groups = [None] + groups
        else:
            groups = [None]
    
        # Disable p-values if group info is missing
        show_p = self.show_p_values and self.group_variable is not None and self.group_labels
    
        # Build column headers
        colnames = ["Variable", "Class"] + ([self.group_labels.get(g, str(g)) if g is not None else "Total" for g in groups])
        if show_p:
            colnames.append("P-value")
    
        # Row for counts
        count_row = ["N", None]
        for g in groups:
            subset = self.df if g is None else self.df[self.df[self.group_variable] == g]
            n = len(subset)
            count_row.append(f"{n:,}" if self.thousands_sep else str(n))
        if show_p:
            count_row.append("")
        rows.append(count_row)
    
        # Variable rows
        for var_cfg in self.variables:
            var = var_cfg["var"]
            name = var_cfg["name"]
            typ = var_cfg["type"]
            p_val = self._compute_p_value(var_cfg) if show_p else None
            p_text = self._format_p(p_val) if p_val is not None else ""
    
            if typ == "categorical":
                class_labels = var_cfg.get("class_labels", {})
                subrows = []
                for cls, label in class_labels.items():
                    row = [None, label]
                    for g in groups:
                        subset = self.df if g is None else self.df[self.df[self.group_variable] == g]
                        summary = self._categorical_summary(subset[var], {cls: label})
                        row.append(summary.get(label, "0 (0.0%)"))
                    if show_p:
                        row.append("")
                    subrows.append(row)
                subrows[0][0] = name
                if show_p:
                    subrows[0][-1] = p_text
                rows.extend(subrows)
    
            elif typ == "continuous":
                stat = var_cfg["stat"]
                decimals = var_cfg.get("decimals", self.percent_decimals)
                row = [name, None]
                for g in groups:
                    subset = self.df if g is None else self.df[self.df[self.group_variable] == g]
                    summary = self._continuous_summary(subset[var], stat, decimals)
                    row.append(summary)
                if show_p:
                    row.append(p_text)
                rows.append(row)
    
        return pd.DataFrame(rows, columns=colnames)

        return pd.DataFrame(rows, columns=colnames)

    def _write_excel(self, table_df: pd.DataFrame, filepath: str):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Demographics"
    
        thin_border = Border(bottom=Side(style='thin', color='000000'))
    
        # Title row (row 1)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=table_df.shape[1])
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = self.table_name
        title_cell.font = Font(bold=True, size=11, name="Times New Roman")
        title_cell.alignment = Alignment(horizontal='left')
    
        # Apply bottom border to entire merged title row
        for c in range(1, table_df.shape[1] + 1):
            ws.cell(row=1, column=c).border = thin_border
    
        # Write data rows starting at row 2
        for r_idx, row in enumerate(dataframe_to_rows(table_df, index=False, header=True), start=2):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.value = value
                cell.font = Font(name="Times New Roman")
                cell.alignment = Alignment(horizontal='left')
    
            # Apply border to header row
            if r_idx == 2:
                for c_idx in range(1, table_df.shape[1] + 1):
                    ws.cell(row=r_idx, column=c_idx).border = thin_border
    
        # Apply border to last data row (r_idx ends at last row)
        for c_idx in range(1, table_df.shape[1] + 1):
            ws.cell(row=r_idx, column=c_idx).border = thin_border
    
        wb.save(filepath)
        print(f"Saved Excel file to: {filepath}")

