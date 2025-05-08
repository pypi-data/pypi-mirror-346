from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import logging

logging.basicConfig(level=logging.DEBUG)


def add_numeric_sheet_to_file(path: str) -> None:
    """Exporta um novo xlsx com todo o conteúdo do arquivo informado no path,
    adicionado um sheet (planilha) contendo apenas as colunas numéricas do 
    arquivo enviado.

    Args:
        path (str): O caminho do arquivo xlsx com o nome a extensão. Exemplo:
            assets/sample.xlsx
    """
    numeric_new_file = path.replace('.xlsx', '_numeric.xlsx')

    wb_original = load_workbook(path)

    wb_original.save(numeric_new_file)
    wb_new = load_workbook(numeric_new_file)

    sheets = wb_original.sheetnames

    for sheet_name in sheets:
        numeric_sheet_name = f"numeric_{sheet_name}"
        if numeric_sheet_name not in wb_new.sheetnames:
            wb_new.create_sheet(numeric_sheet_name)
        
        old_sheet = wb_original[sheet_name]
        numeric_sheet = wb_new[numeric_sheet_name]

        for col in old_sheet.iter_cols():
            if all(isinstance(cell.value, (int, float)) or cell.row == 1 for cell in col):
                for cell in col:
                    numeric_sheet.cell(row=cell.row, column=cell.column, value=cell.value)

    wb_new.save(numeric_new_file)

    for col in reversed(list(numeric_sheet.iter_cols(min_row=1, max_row=numeric_sheet.max_row, min_col=1, max_col=numeric_sheet.max_column))):
        if all(cell.value is None for cell in col):
            numeric_sheet.delete_cols(col[0].column)

    wb_new.save(numeric_new_file)


def gera_metrica(path: str):
    """Gera as métricas (Média, Máximo, Mínimo e Somatória) dos campos numéricos presentes 
    no arquivo xlsx fornecido no path. 

    Args:
        path (str): O caminho do arquivo xlsx com o nome a extensão. Exemplo:
            assets/sample.xlsx
    """
    numeric_path = path.replace('.xlsx', '_numeric.xlsx')
    report_path = path.replace(".xlsx", "_analise_quantitativa.xlsx")
    wb_original = load_workbook(numeric_path)
    
    wb_original.save(report_path)
    wb_new = load_workbook(report_path)

    wb_new.create_sheet("AnaliseQuantitativa")
    wb_new.save(report_path)

    sheets: list = wb_original.sheetnames

    for sheet in sheets:
        if "numeric_" in sheet:
            current_sheet = sheet
    
    col_names = [cel.value for cel in wb_original[current_sheet][1]]
    wb_analise = wb_new["AnaliseQuantitativa"]

    for col_index, name in enumerate(col_names, start=1):
        col_letter = get_column_letter(col_index)
        logging.info("Gerando a Somatória")

        #definindo largura da coluna
        wb_analise.column_dimensions[f"{col_letter}"].width = 55
        
        #definindo fonte da linha
        wb_analise[f"{col_letter}1"].font = Font(name='Calibri', size=16, bold=True, color='0000FF')
        
        #definindo a altura da linha
        wb_analise.row_dimensions[1].height = 30 
        wb_analise[f"{col_letter}1"] = f"Somatória dos valores de: {name}"

        wb_analise[f"{col_letter}2"] = f"=SUM({sheet}!{col_letter}:{col_letter})"
        logging.info("Gerando a média")
        #definindo fonte da linha
        wb_analise[f"{col_letter}5"].font = Font(name='Calibri', size=16, bold=True, color='0000FF')
        
        wb_analise.row_dimensions[5].height = 30
        wb_analise[f"{col_letter}5"] = f"Média dos valores de: {name}"
        
        wb_analise[f"{col_letter}6"] = f"=AVERAGE({sheet}!{col_letter}:{col_letter})"
        logging.info("Gerando o máximo")
        
        #definindo fonte da linha
        wb_analise[f"{col_letter}9"].font = Font(name='Calibri', size=16, bold=True, color='0000FF')
        wb_analise.row_dimensions[9].height = 30
        wb_analise[f"{col_letter}9"] = f"Máximo dos valores de: {name}"

        wb_analise[f"{col_letter}10"] = f"=MAX({sheet}!{col_letter}:{col_letter})"
        
        logging.info("Gerando o minimo")
        #definindo fonte da linha
        wb_analise[f"{col_letter}13"].font = Font(name='Calibri', size=16, bold=True, color='0000FF')
        wb_analise.row_dimensions[13].height = 30
        wb_analise[f"{col_letter}13"] = f"Mínimo dos valores de: {name}"

        wb_analise[f"{col_letter}14"] = f"=MIN({sheet}!{col_letter}:{col_letter})"
    
    wb_new.save(report_path)
    logging.info("Relatório gerado com sucesso.")


if __name__ == "__main__":
    sample_file = "assets/sample.xlsx"
    add_numeric_sheet_to_file(sample_file)
    gera_metrica(sample_file)