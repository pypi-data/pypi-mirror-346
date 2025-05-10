"""
Módulo para geração de relatórios de análise de associação de produtos.

Este módulo contém funções para criar relatórios detalhados em Excel
com os resultados da análise de associação de produtos.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, Reference


def criar_excel_detalhado(associacoes_por_produto, contagem_produtos, output_file="relatorio_associacoes_completo.xlsx", max_produtos=20):
    """
    Cria um relatório Excel detalhado com os resultados da análise de associação de produtos.
    
    Args:
        associacoes_por_produto (dict): Dicionário com os rankings de associações para cada produto
        contagem_produtos (pandas.DataFrame): DataFrame com a contagem de frequência de cada produto
        output_file (str, opcional): Nome do arquivo Excel de saída. Padrão é "relatorio_associacoes_completo.xlsx".
        max_produtos (int, opcional): Número máximo de produtos para criar abas individuais. Padrão é 20.
        
    Returns:
        str: Caminho do arquivo Excel criado
    """
    # Criar um novo workbook
    wb = Workbook()
    
    # Definir estilos
    titulo_font = Font(name='Arial', size=14, bold=True, color="FFFFFF")
    cabecalho_font = Font(name='Arial', size=12, bold=True)
    dados_font = Font(name='Arial', size=11)
    
    titulo_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    cabecalho_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    
    titulo_alignment = Alignment(horizontal='center', vertical='center')
    cabecalho_alignment = Alignment(horizontal='center', vertical='center')
    dados_alignment = Alignment(horizontal='left', vertical='center')
    
    borda = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # 1. Criar planilha de resumo
    ws_resumo = wb.active
    ws_resumo.title = "Resumo"
    
    # Título
    ws_resumo.merge_cells('A1:C1')
    ws_resumo['A1'] = "Análise de Associação de Produtos - Resumo"
    ws_resumo['A1'].font = titulo_font
    ws_resumo['A1'].fill = titulo_fill
    ws_resumo['A1'].alignment = titulo_alignment
    
    # Cabeçalho
    ws_resumo['A3'] = "Produto"
    ws_resumo['B3'] = "Frequência"
    ws_resumo['C3'] = "% do Total"
    
    for cell in ws_resumo['A3:C3'][0]:
        cell.font = cabecalho_font
        cell.fill = cabecalho_fill
        cell.alignment = cabecalho_alignment
        cell.border = borda
    
    # Dados
    total_frequencia = contagem_produtos['Frequência'].sum()
    
    for i, (_, row) in enumerate(contagem_produtos.iterrows(), 4):
        ws_resumo[f'A{i}'] = row['Produto']
        ws_resumo[f'B{i}'] = row['Frequência']
        ws_resumo[f'C{i}'] = f"{(row['Frequência'] / total_frequencia) * 100:.2f}%"
        
        for col in ['A', 'B', 'C']:
            ws_resumo[f'{col}{i}'].font = dados_font
            ws_resumo[f'{col}{i}'].alignment = dados_alignment
            ws_resumo[f'{col}{i}'].border = borda
    
    # Ajustar largura das colunas
    ws_resumo.column_dimensions['A'].width = 50
    ws_resumo.column_dimensions['B'].width = 15
    ws_resumo.column_dimensions['C'].width = 15
    
    # 2. Criar planilha para cada produto
    # Ordenar produtos por frequência
    produtos_ordenados = contagem_produtos.sort_values('Frequência', ascending=False)
    
    # Determinar quantos produtos incluir (limite para não criar muitas abas)
    max_produtos = min(max_produtos, len(produtos_ordenados))
    print(f"Criando {max_produtos} abas para os produtos mais frequentes...")
    
    for idx, (_, row) in enumerate(produtos_ordenados.head(max_produtos).iterrows()):
        produto = row['Produto']
        freq = row['Frequência']
        
        # Criar nome seguro para a planilha (limitar comprimento, remover caracteres especiais)
        nome_planilha = f"Prod_{idx+1}"
        
        # Criar planilha
        ws_produto = wb.create_sheet(nome_planilha)
        
        # Título
        ws_produto.merge_cells('A1:C1')
        ws_produto['A1'] = f"Associações do Produto: {produto}"
        ws_produto['A1'].font = titulo_font
        ws_produto['A1'].fill = titulo_fill
        ws_produto['A1'].alignment = titulo_alignment
        
        # Informações do produto
        ws_produto['A3'] = "Produto:"
        ws_produto['B3'] = produto
        ws_produto['A4'] = "Frequência Total:"
        ws_produto['B4'] = freq
        
        # Cabeçalho da tabela de associações
        ws_produto['A6'] = "Produto Associado"
        ws_produto['B6'] = "Frequência"
        ws_produto['C6'] = "% das Compras"
        
        for cell in ws_produto['A6:C6'][0]:
            cell.font = cabecalho_font
            cell.fill = cabecalho_fill
            cell.alignment = cabecalho_alignment
            cell.border = borda
        
        # Dados de associação
        if produto in associacoes_por_produto:
            associacoes = associacoes_por_produto[produto]
            
            for i, (_, assoc_row) in enumerate(associacoes.iterrows(), 7):
                ws_produto[f'A{i}'] = assoc_row['Produto_Associado']
                ws_produto[f'B{i}'] = assoc_row['Frequência']
                ws_produto[f'C{i}'] = f"{(assoc_row['Frequência'] / freq) * 100:.2f}%"
                
                for col in ['A', 'B', 'C']:
                    ws_produto[f'{col}{i}'].font = dados_font
                    ws_produto[f'{col}{i}'].alignment = dados_alignment
                    ws_produto[f'{col}{i}'].border = borda
            
            # Adicionar gráfico
            if len(associacoes) > 0:
                # Limitar ao top 10 para o gráfico
                top_n = min(10, len(associacoes))
                
                chart = BarChart()
                chart.type = "col"
                chart.style = 10
                chart.title = "Top Produtos Associados"
                chart.y_axis.title = "Frequência"
                chart.x_axis.title = "Produtos"
                
                # Definir dados para o gráfico
                data = Reference(ws_produto, min_col=2, min_row=6, max_row=6+top_n, max_col=2)
                cats = Reference(ws_produto, min_col=1, min_row=7, max_row=6+top_n)
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(cats)
                chart.shape = 4
                
                # Adicionar gráfico
                ws_produto.add_chart(chart, f"E7")
        
        # Ajustar largura das colunas
        ws_produto.column_dimensions['A'].width = 50
        ws_produto.column_dimensions['B'].width = 15
        ws_produto.column_dimensions['C'].width = 15
    
    # 3. Criar planilha com todos os rankings completos
    ws_rankings = wb.create_sheet("Rankings Completos")
    
    # Título
    ws_rankings.merge_cells('A1:E1')
    ws_rankings['A1'] = "Ranking Completo de Associações por Produto"
    ws_rankings['A1'].font = titulo_font
    ws_rankings['A1'].fill = titulo_fill
    ws_rankings['A1'].alignment = titulo_alignment
    
    # Cabeçalho
    ws_rankings['A3'] = "Produto Principal"
    ws_rankings['B3'] = "Ranking"
    ws_rankings['C3'] = "Produto Associado"
    ws_rankings['D3'] = "Frequência"
    ws_rankings['E3'] = "% das Compras"
    
    for cell in ws_rankings['A3:E3'][0]:
        cell.font = cabecalho_font
        cell.fill = cabecalho_fill
        cell.alignment = cabecalho_alignment
        cell.border = borda
    
    # Adicionar dados de todos os produtos
    linha = 4
    
    # Ordenar produtos por frequência antes de adicionar ao ranking completo
    for produto in produtos_ordenados['Produto']:
        # Se não houver associações para este produto, continuar
        if produto not in associacoes_por_produto or associacoes_por_produto[produto].empty:
            continue
        
        freq_produto = contagem_produtos[contagem_produtos['Produto'] == produto]['Frequência'].values[0]
        associacoes = associacoes_por_produto[produto]
        
        # Para cada produto associado
        for i, (_, assoc_row) in enumerate(associacoes.iterrows()):
            ws_rankings[f'A{linha}'] = produto
            ws_rankings[f'B{linha}'] = i + 1  # Ranking
            ws_rankings[f'C{linha}'] = assoc_row['Produto_Associado']
            ws_rankings[f'D{linha}'] = assoc_row['Frequência']
            ws_rankings[f'E{linha}'] = f"{(assoc_row['Frequência'] / freq_produto) * 100:.2f}%"
            
            for col in ['A', 'B', 'C', 'D', 'E']:
                ws_rankings[f'{col}{linha}'].font = dados_font
                ws_rankings[f'{col}{linha}'].alignment = dados_alignment
                ws_rankings[f'{col}{linha}'].border = borda
            
            linha += 1
        
        # Adicionar uma linha em branco entre produtos
        linha += 1
    
    # Ajustar largura das colunas
    ws_rankings.column_dimensions['A'].width = 50
    ws_rankings.column_dimensions['B'].width = 10
    ws_rankings.column_dimensions['C'].width = 50
    ws_rankings.column_dimensions['D'].width = 15
    ws_rankings.column_dimensions['E'].width = 15
    
    # 4. Criar planilha de insights
    ws_insights = wb.create_sheet("Insights e Recomendações")
    
    # Título
    ws_insights.merge_cells('A1:B1')
    ws_insights['A1'] = "Insights e Recomendações de Marketing"
    ws_insights['A1'].font = titulo_font
    ws_insights['A1'].fill = titulo_fill
    ws_insights['A1'].alignment = titulo_alignment
    
    # Top produtos para insights
    top_produtos = produtos_ordenados.head(3)['Produto'].tolist()
    top_produtos_str = ", ".join(top_produtos)
    
    # Adicionar insights
    insights = [
        {"titulo": "1. Produtos Mais Vendidos", 
         "texto": f"Os produtos mais populares são {top_produtos_str}. "
                  "Considere destacá-los na loja física e online para aumentar ainda mais as vendas."},
        
        {"titulo": "2. Oportunidades de Cross-Selling", 
         "texto": "Para cada produto, consulte sua aba específica ou a planilha 'Rankings Completos' "
                  "para identificar quais produtos são mais frequentemente comprados junto. "
                  "Use estas informações para criar recomendações personalizadas para os clientes."},
        
        {"titulo": "3. Disposição na Loja", 
         "texto": "Posicione produtos que são frequentemente comprados juntos em locais próximos na loja, "
                  "facilitando a visualização pelo cliente. Consulte as abas de produtos específicos "
                  "para determinar o posicionamento ideal."},
        
        {"titulo": "4. Pacotes Promocionais", 
         "texto": "Crie pacotes promocionais com produtos que são frequentemente comprados juntos, "
                  "oferecendo um desconto na compra conjunta. Isto pode aumentar o valor médio do pedido."},
        
        {"titulo": "5. Marketing Direcionado", 
         "texto": "Utilize as informações de associação para criar campanhas de marketing direcionadas. "
                  "Por exemplo, envie promoções de produtos associados para clientes que compraram um produto específico."},
    ]
    
    row = 3
    for insight in insights:
        # Título do insight
        ws_insights.merge_cells(f'A{row}:B{row}')
        ws_insights[f'A{row}'] = insight["titulo"]
        ws_insights[f'A{row}'].font = cabecalho_font
        
        row += 1
        
        # Texto do insight
        ws_insights.merge_cells(f'A{row}:B{row}')
        ws_insights[f'A{row}'] = insight["texto"]
        ws_insights[f'A{row}'].font = dados_font
        ws_insights[f'A{row}'].alignment = Alignment(wrap_text=True)
        
        row += 2
    
    # Adicionar recomendações específicas baseadas nos dados
    ws_insights.merge_cells(f'A{row}:B{row}')
    ws_insights[f'A{row}'] = "Recomendações Específicas por Produto:"
    ws_insights[f'A{row}'].font = cabecalho_font
    
    row += 1
    
    # Adicionar recomendações específicas para os top 5 produtos
    for produto in produtos_ordenados.head(5)['Produto']:
        if produto in associacoes_por_produto and not associacoes_por_produto[produto].empty:
            associacoes = associacoes_por_produto[produto].head(3)
            
            if not associacoes.empty:
                produtos_associados = associacoes['Produto_Associado'].tolist()
                produtos_associados_str = ", ".join(produtos_associados)
                
                recomendacao = f"Para clientes que compram {produto}, recomende: {produtos_associados_str}"
                
                ws_insights.merge_cells(f'A{row}:B{row}')
                ws_insights[f'A{row}'] = recomendacao
                ws_insights[f'A{row}'].font = dados_font
                ws_insights[f'A{row}'].alignment = Alignment(wrap_text=True)
                
                row += 1
    
    # Ajustar largura das colunas
    ws_insights.column_dimensions['A'].width = 15
    ws_insights.column_dimensions['B'].width = 65
    
    # Salvar o arquivo
    wb.save(output_file)
    print(f"Arquivo Excel salvo como {output_file}")
    
    return output_file