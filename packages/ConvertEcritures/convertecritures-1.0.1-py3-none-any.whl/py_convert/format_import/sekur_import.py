import polars as pl
import dateparser
from openpyxl import load_workbook
from .base_import import BaseImport
from ..error import run_error

class SekurImport(BaseImport):
    """Gestion d'import d'un fichier Excel au format SEKUR."""
    
    def name(self):
        return "SEKUR"
    
    def validate_format(self):
        if self.path.suffix.lower() != ".xlsx":
            run_error(f"Le format {self.name()} nécessite un fichier .xlsx")
            return False
        return True
    
    def process_file(self):
        # Importe le fichier excel en dataframe avec l'en-tête en ligne 4
        df = pl.read_excel(source=self.path, read_options={"header_row": 3})

        # Permet de supprimer les dernières lignes de totaux
        df = df.filter(pl.col("Jour").cast(pl.Utf8).str.contains(r'^\d+$'))

        if df.columns[0] != "Jour": 
            run_error("Ce fichier excel n'est pas au format du logiciel de vente SEKUR")
            return None

        # Récupère le mois et l'année des écritures du journal de vente
        wb = load_workbook(filename=self.path)
        ws = wb.worksheets[0]
        valeur = ws["F3"].value
        date_obj = dateparser.parse(valeur)

        # Transforme la colonne jour en colonne date
        df = df.with_columns(
            (pl.col("Jour").cast(pl.Utf8) + "-" + date_obj.strftime('%m-%Y'))
            .str.strptime(pl.Date, "%d-%m-%Y")
            .alias("EcritureDate"))
        df = df.drop("Jour")

        # Remplie les valeurs manquantes de ces 3 colonnes
        df = df.with_columns(
            pl.col("Compte").fill_null("CDIVERS"),
            pl.col("Débit").fill_null(0.0),
            pl.col("Crédit").fill_null(0.0),
            )

        # Vérifie si des factures d'avoirs doivent être inversées
        df = df.with_columns(
            pl.when(
                (pl.col("N° Facture").str.slice(0, 1) == "A") &
                (pl.col("Compte").str.starts_with("C")) & 
                (pl.col("Débit") > 1e-3)
            )
            .then(pl.lit(True))
            .otherwise(pl.lit(False))
            .alias("Condition")
        )

        # Inverse le sens des factures d'avoirs lorsque c'est nécessaire
        df = df.with_columns((pl
                            .when(pl.col("Condition").over("N° Facture").any())
                            .then(pl.col("Crédit"))
                            .otherwise(pl.col("Débit"))
                            .alias("Debit")),
                            (pl
                            .when(pl.col("Condition").over("N° Facture").any())
                            .then(pl.col("Débit"))
                            .otherwise(pl.col("Crédit"))
                            .alias("Credit"))
                            )
        df = df.drop("Débit", "Crédit", "Condition")

        # Création des colonnes manquantes
        df = df.with_columns(
            pl.lit("VE").alias("JournalCode").cast(pl.String),
            pl.lit(None).alias("JournalLib").cast(pl.String),
            pl.lit(None).alias("EcritureNum").cast(pl.String),
            pl.lit(None).alias("CompAuxNum").cast(pl.String),
            pl.lit(None).alias("CompAuxLib").cast(pl.String),
            pl.col("EcritureDate").alias("PieceDate").cast(pl.Date),
            pl.lit(None).alias("EcritureLet").cast(pl.String),
            pl.lit(None).alias("DateLet").cast(pl.Date),
            pl.lit(None).alias("ValidDate").cast(pl.Date),
            pl.lit(None).alias("Montantdevise").cast(pl.Float64),
            pl.lit(None).alias("Idevise").cast(pl.String),
            pl.lit(None).alias("EcheanceDate").cast(pl.Date)
            )

        # Réorganise l'ordre des colonnes
        df = df.select("JournalCode",
                    "JournalLib",
                    "EcritureNum",
                    "EcritureDate",
                    pl.col("Compte").alias("CompteNum"),
                    pl.col("Intitulé").alias("CompteLib"),
                    "CompAuxNum",
                    "CompAuxLib",
                    pl.col("N° Facture").alias("PieceRef"),
                    "PieceDate",
                    pl.col("Libellé de l'écriture").alias("EcritureLib"),
                    "Debit",
                    "Credit",
                    "EcritureLet",
                    "DateLet",
                    "ValidDate",
                    "Montantdevise",
                    "Idevise",
                    "EcheanceDate"
                    )
        
        return df
    