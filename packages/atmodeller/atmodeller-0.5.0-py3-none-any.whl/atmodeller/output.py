#
# Copyright 2024 Dan J. Bower
#
# This file is part of Atmodeller.
#
# Atmodeller is free software: you can redistribute it and/or modify it under the terms of the GNU
# General Public License as published by the Free Software Foundation, either version 3 of the
# License, or (at your option) any later version.
#
# Atmodeller is distributed in the hope that it will be useful, but WITHOUT ANY WARRANTY; without
# even the implied warranty of MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the GNU
# General Public License for more details.
#
# You should have received a copy of the GNU General Public License along with Atmodeller. If not,
# see <https://www.gnu.org/licenses/>.
#
"""Output

This uses existing functions as much as possible to calculate desired output quantities, where some
must be vmapped to compute the output.
"""

from __future__ import annotations

import logging
import pickle
from pathlib import Path
from typing import Any, Callable

import equinox as eqx
import jax
import jax.numpy as jnp
import numpy as np
import numpy.typing as npt
import pandas as pd
from jax import Array
from jax.tree_util import tree_map
from jax.typing import ArrayLike
from molmass import Formula
from openpyxl.styles import PatternFill
from scipy.constants import mega

from atmodeller.constants import AVOGADRO
from atmodeller.containers import (
    FixedParameters,
    Planet,
    Species,
    SpeciesCollection,
    TracedParameters,
)
from atmodeller.engine import (
    get_atmosphere_log_molar_mass,
    get_atmosphere_log_volume,
    get_element_density,
    get_element_density_in_melt,
    get_log_activity,
    get_log_number_density_from_log_pressure,
    get_pressure_from_log_number_density,
    get_species_density_in_melt,
    get_species_ppmw_in_melt,
    get_total_pressure,
    objective_function,
)
from atmodeller.interfaces import RedoxBufferProtocol
from atmodeller.thermodata import IronWustiteBuffer
from atmodeller.utilities import unit_conversion, vmap_axes_spec

logger: logging.Logger = logging.getLogger(__name__)


class Output:
    """Output

    Args:
        species: Species
        solution: Array output from solve
        solver_status: Solver status
        solver_steps: Number of solver steps
        fixed_parameters: Fixed parameters
        traced_parameters: Traced parameters
    """

    def __init__(
        self,
        species: SpeciesCollection,
        solution: Array,
        solver_status: Array,
        solver_steps: Array,
        fixed_parameters: FixedParameters,
        traced_parameters: TracedParameters,
    ):
        logger.debug("Creating Output")
        self._species: SpeciesCollection = species
        self._solution: Array = solution
        self._solver_status: Array = solver_status
        self._solver_steps: Array = solver_steps
        self._fixed_parameters: FixedParameters = fixed_parameters
        self._traced_parameters: TracedParameters = traced_parameters

        # Calculate the index at which to split the array
        split_index: int = self._solution.shape[1] - self._species.number_of_stability()
        log_number_density, log_stability = jnp.split(self._solution, [split_index], axis=1)
        # Number of entries in log_number_density is always the total number of species
        self._log_number_density: Array = log_number_density
        # Number of entries in log_stability could be between 0 and the total number of species
        self._log_stability: Array = log_stability
        # Caching output to avoid recomputation
        self._cached_dict: dict[str, dict[str, Array]] | None = None
        self._cached_dataframes: dict[str, pd.DataFrame] | None = None

    @property
    def condensed_species_indices(self) -> Array:
        """Condensed species indices"""
        return self._species.get_condensed_species_indices()

    @property
    def gas_species_indices(self) -> Array:
        """Gas species indices"""
        return self._fixed_parameters.gas_species_indices

    @property
    def log_number_density(self) -> Array:
        """Log number density"""
        return self._log_number_density

    @property
    def log_stability(self) -> Array:
        """Log stability of relevant species"""
        return self._log_stability

    @property
    def molar_mass(self) -> Array:
        """Molar mass of all species"""
        return self._fixed_parameters.molar_masses

    @property
    def number_solutions(self) -> int:
        """Number of solutions"""
        return self.log_number_density.shape[0]

    @property
    def planet(self) -> Planet:
        """Planet"""
        return self._traced_parameters.planet

    @property
    def stability_species_mask(self) -> Array:
        """Stability species mask"""
        return self._species.get_stability_species_mask()

    @property
    def temperature(self) -> Array:
        """Temperature"""
        return jnp.asarray(self.planet.temperature)

    @property
    def temperature_vmap(self) -> int | None:
        """Axis for temperature vmap"""
        return vmap_axes_spec(self._traced_parameters.planet).temperature

    @property
    def traced_parameters_vmap(self) -> TracedParameters:
        """Axis for traced parameters vmap"""
        return vmap_axes_spec(self._traced_parameters)

    def activity(self) -> Array:
        """Gets the activity of all species

        Returns:
            Activity of all species
        """
        return jnp.exp(self.log_activity())

    def activity_without_stability(self) -> Array:
        """Gets the activity without stability of all species

        Returns:
            Activity without stability of all species
        """
        return jnp.exp(self.log_activity_without_stability())

    def asdict(self) -> dict[str, dict[str, Array]]:
        """All output in a dictionary, with caching.

        TODO: This function is slow and is an obvious candidate for speeding up.

        Returns:
            Dictionary of all output
        """
        if self._cached_dict is not None:
            logger.info("Returning cached asdict output")
            return self._cached_dict  # Return cached result

        logger.info("Computing asdict output")
        fugacity_matrix: Array = jnp.array(self._fixed_parameters.fugacity_matrix)
        formula_matrix_constraints: Array = jnp.array(
            self._fixed_parameters.formula_matrix_constraints
        )

        out: dict[str, dict[str, Array]] = {}
        out |= self.condensed_species_asdict()
        out |= self.gas_species_asdict()
        out |= self.elements_asdict()

        out["planet"] = expand_dict(self.planet.asdict(), self.number_solutions)
        out["atmosphere"] = self.atmosphere_asdict()
        # temperature and pressure have already been expanded to the number of solutions
        temperature: Array = out["planet"]["surface_temperature"]
        pressure: Array = out["atmosphere"]["pressure"]
        # Convenient to also attach temperature to the atmosphere output
        out["atmosphere"]["temperature"] = temperature
        out["raw_solution"] = self.raw_solution_asdict()

        out["constraints"] = {}
        if formula_matrix_constraints.size > 0:
            out["constraints"] |= expand_dict(
                self._traced_parameters.mass_constraints.asdict(), self.number_solutions
            )
        if fugacity_matrix.size > 0:
            out["constraints"] |= self._traced_parameters.fugacity_constraints.asdict(
                temperature, pressure
            )
        out["residual"] = self.residual_asdict()  # type: ignore since uses int for keys

        if "O2_g" in out:
            logger.debug("Found O2_g so back-computing log10 shift for fO2")
            log10_fugacity: Array = jnp.log10(out["O2_g"]["fugacity"])
            temperature: Array = out["planet"]["surface_temperature"]
            pressure: Array = out["atmosphere"]["pressure"]
            buffer: RedoxBufferProtocol = IronWustiteBuffer()
            # Shift at 1 bar
            buffer_at_one_bar: ArrayLike = buffer.log10_fugacity(temperature, 1.0)
            log10_shift_at_one_bar: Array = log10_fugacity - buffer_at_one_bar
            logger.debug("log10_shift_at_1bar = %s", log10_shift_at_one_bar)
            out["O2_g"]["log10dIW_1_bar"] = log10_shift_at_one_bar
            # Shift at actual pressure
            buffer_at_P: ArrayLike = buffer.log10_fugacity(temperature, pressure)
            log10_shift_at_P: Array = log10_fugacity - buffer_at_P
            logger.debug("log10_shift_at_P = %s", log10_shift_at_P)
            out["O2_g"]["log10dIW_P"] = log10_shift_at_P

        out["solver"] = {"status": self._solver_status, "steps": self._solver_steps}

        # Convert all arrays in the dictionary to numpy arrays. Using the same functions that the
        # engine uses makes sense to avoid duplication and ensure consistency, but for the output
        # we can convert all arrays to numpy. Then the output will only consistent of dictionaries
        # and numpy arrays, which might simplify the pickling/unpickling process for end-users who
        # only care about the results and do not necessary want to use JAX types to process
        # the output. To this point the arrays are of type
        # <class 'jaxlib.xla_extension.ArrayImpl'>.
        # TODO: could maybe use numpy method?
        def convert_to_numpy(d) -> None:
            for key, value in d.items():
                if isinstance(value, dict):
                    convert_to_numpy(value)
                else:
                    logger.debug("Array type before conversion = %s", type(d[key]))
                    d[key] = np.array(value)
                    logger.debug("Array type after conversion = %s", type(d[key]))

        logger.debug("Convert all arrays to numpy")
        convert_to_numpy(out)

        self._cached_dict = out  # Cache result

        return out

    def atmosphere_asdict(self) -> dict[str, Array]:
        """Gets the atmosphere properties

        Returns:
            Atmosphere properties
        """
        out: dict[str, Array] = {}

        log_number_density_from_log_pressure_func: Callable = eqx.filter_vmap(
            get_log_number_density_from_log_pressure, in_axes=(0, self.temperature_vmap)
        )
        log_number_density = log_number_density_from_log_pressure_func(
            jnp.log(self.total_pressure()), self.temperature
        )
        # Must be 2-D to align arrays for computing number-density-related quantities
        number_density: Array = jnp.exp(log_number_density)[:, jnp.newaxis]
        molar_mass: Array = self.atmosphere_molar_mass()[:, jnp.newaxis]
        out: dict[str, Array] = self._get_number_density_output(
            number_density, molar_mass, "species_"
        )
        # Species mass is simply mass, so rename for clarity
        out["mass"] = out.pop("species_mass")

        out["molar_mass"] = molar_mass
        # Ensure all arrays are 1-D, which is required for creating dataframes
        out = {key: value.ravel() for key, value in out.items()}

        out["pressure"] = self.total_pressure()
        out["volume"] = self.atmosphere_volume()
        out["element_number_density"] = jnp.sum(self.element_density_gas(), axis=1)
        out["element_number"] = out["element_number_density"] * out["volume"]
        out["element_moles"] = out["element_number"] / AVOGADRO

        return out

    def atmosphere_log_molar_mass(self) -> Array:
        """Gets log molar mass of the atmosphere

        Returns:
            Log molar mass of the atmosphere
        """
        atmosphere_log_molar_mass_func: Callable = eqx.filter_vmap(
            get_atmosphere_log_molar_mass, in_axes=(None, 0)
        )
        atmosphere_log_molar_mass: Array = atmosphere_log_molar_mass_func(
            self._fixed_parameters, self.log_number_density
        )

        return atmosphere_log_molar_mass

    def atmosphere_molar_mass(self) -> Array:
        """Gets the molar mass of the atmosphere

        Returns:
            Molar mass of the atmosphere
        """
        return jnp.exp(self.atmosphere_log_molar_mass())

    def atmosphere_log_volume(self) -> Array:
        """Gets the log volume of the atmosphere

        Returns:
            Log volume of the atmosphere
        """
        atmosphere_log_volume_func: Callable = eqx.filter_vmap(
            get_atmosphere_log_volume,
            in_axes=(
                None,
                0,
                vmap_axes_spec(self._traced_parameters.planet),
            ),
        )
        atmosphere_log_volume: Array = atmosphere_log_volume_func(
            self._fixed_parameters,
            self.log_number_density,
            self.planet,
        )

        return atmosphere_log_volume

    def atmosphere_volume(self) -> Array:
        """Gets the volume of the atmosphere

        Returns:
            Volume of the atmosphere
        """
        return jnp.exp(self.atmosphere_log_volume())

    def total_pressure(self) -> Array:
        """Gets total pressure

        Returns:
            Total pressure
        """
        total_pressure_func: Callable = eqx.filter_vmap(
            get_total_pressure, in_axes=(None, 0, self.temperature_vmap)
        )
        total_pressure: Array = total_pressure_func(
            self._fixed_parameters, self.log_number_density, self.temperature
        )

        return total_pressure

    def condensed_species_asdict(self) -> dict[str, dict[str, Array]]:
        """Gets the condensed species output as a dictionary

        Returns:
            Condensed species output as a dictionary
        """
        molar_mass: Array = self.species_molar_mass_expanded()
        number_density_condensed: Array = jnp.take(
            self.number_density(), self.condensed_species_indices, axis=1
        )
        activity_condensed: Array = jnp.take(
            self.activity(), self.condensed_species_indices, axis=1
        )
        molar_mass_condensed: Array = jnp.take(molar_mass, self.condensed_species_indices, axis=1)
        condensed_species: tuple[Species, ...] = tuple(
            self._species[ii] for ii in self.condensed_species_indices
        )
        out: dict[str, Array] = self._get_number_density_output(
            number_density_condensed, molar_mass_condensed, "total_"
        )
        out["molar_mass"] = molar_mass_condensed
        out["activity"] = activity_condensed

        split_dict: list[dict[str, Array]] = split_dict_by_columns(out)
        species_out: dict[str, dict[str, Array]] = {
            species.name: split_dict[ii] for ii, species in enumerate(condensed_species)
        }

        return species_out

    def elements_asdict(self) -> dict[str, dict[str, Array]]:
        """Gets the element properties as a dictionary

        Returns:
            Element outputs as a dictionary
        """
        molar_mass: Array = self.element_molar_mass_expanded()
        atmosphere: Array = self.element_density_gas()
        condensed: Array = self.element_density_condensed()
        dissolved: Array = self.element_density_dissolved()
        total: Array = atmosphere + condensed + dissolved

        out: dict[str, Array] = self._get_number_density_output(
            atmosphere, molar_mass, "atmosphere_"
        )
        out |= self._get_number_density_output(condensed, molar_mass, "condensed_")
        out |= self._get_number_density_output(dissolved, molar_mass, "dissolved_")
        out |= self._get_number_density_output(total, molar_mass, "total_")

        out["molar_mass"] = molar_mass
        out["degree_of_condensation"] = out["condensed_number"] / out["total_number"]
        out["volume_mixing_ratio"] = out["atmosphere_number"] / jnp.sum(
            out["atmosphere_number"], axis=1, keepdims=True
        )
        out["atmosphere_ppm"] = out["volume_mixing_ratio"] * mega
        out["atmosphere_ppmw"] = (
            out["atmosphere_mass"] / jnp.sum(out["atmosphere_mass"], axis=1, keepdims=True) * mega
        )

        unique_elements: tuple[str, ...] = self._species.get_unique_elements_in_species()
        if "H" in unique_elements:
            index: int = unique_elements.index("H")
            H_total_moles: Array = out["total_moles"][:, index]
            out["logarithmic_abundance"] = (
                jnp.log10(out["total_moles"] / H_total_moles[:, jnp.newaxis]) + 12
            )

        logger.debug("out = %s", out)

        split_dict: list[dict[str, Array]] = split_dict_by_columns(out)
        logger.debug("split_dict = %s", split_dict)

        elements_out: dict[str, dict[str, Array]] = {
            f"element_{element}": split_dict[ii] for ii, element in enumerate(unique_elements)
        }
        logger.debug("elements_out = %s", elements_out)

        return elements_out

    def element_density_condensed(self) -> Array:
        """Gets the number density of elements in the condensed phase

        Unlike for the objective function, we want the number density of all elements, regardless
        of whether they were used to impose a mass constraint on the system.

        Returns:
            Number density of elements in the condensed phase
        """
        element_density_func: Callable = eqx.filter_vmap(get_element_density, in_axes=(None, 0))
        element_density: Array = element_density_func(
            self.formula_matrix_condensed(), self.log_number_density
        )

        return element_density

    def element_density_dissolved(self) -> Array:
        """Gets the number density of elements dissolved in melt due to species solubility

        Unlike for the objective function, we want the number density of all elements, regardless
        of whether they were used to impose a mass constraint on the system.

        Returns:
            Number density of elements dissolved in melt due to species solubility
        """
        element_density_dissolved_func: Callable = eqx.filter_vmap(
            get_element_density_in_melt,
            in_axes=(self.traced_parameters_vmap, None, None, 0, 0, 0),
        )
        element_density_dissolved: Array = element_density_dissolved_func(
            self._traced_parameters,
            self._fixed_parameters,
            jnp.array(self._fixed_parameters.formula_matrix),
            self.log_number_density,
            self.log_activity(),
            self.atmosphere_log_volume(),
        )

        return element_density_dissolved

    def element_density_gas(self) -> Array:
        """Gets the number density of elements in the gas phase

        Unlike for the objective function, we want the number density of all elements, regardless
        of whether they were used to impose a mass constraint on the system.

        Returns:
            Number density of elements in the gas phase
        """
        element_density_func: Callable = eqx.filter_vmap(get_element_density, in_axes=(None, 0))
        element_density: Array = element_density_func(
            self.formula_matrix_gas(), self.log_number_density
        )

        return element_density

    def element_molar_mass_expanded(self) -> Array:
        """Gets molar mass of elements

        Returns:
            Molar mass of elements
        """
        unique_elements: tuple[str, ...] = self._species.get_unique_elements_in_species()
        molar_mass: Array = jnp.array([Formula(element).mass for element in unique_elements])
        molar_mass = unit_conversion.g_to_kg * molar_mass

        return jnp.tile(molar_mass, (self.number_solutions, 1))

    def _get_modified_formula_matrix(self, indices: Array) -> Array:
        """Gets a modified formula matrix with columns not in indices set to zero.

        Args:
            indices: Indices of columns to keep, where others are set to zero.

        Returns:
            Modified formula matrix
        """
        formula_matrix: Array = jnp.array(self._fixed_parameters.formula_matrix)
        mask: Array = jnp.zeros_like(formula_matrix, dtype=bool)
        mask = mask.at[:, indices].set(True)
        logger.debug("modified_formula_matrix = %s", mask)
        modified_formula_matrix: Array = formula_matrix * mask

        return modified_formula_matrix

    def formula_matrix_condensed(self) -> Array:
        """Formula matrix for condensed species

        Only columns in the formula matrix referring to condensed species have non-zero values.

        Returns:
            Formula matrix for condensed species
        """
        return self._get_modified_formula_matrix(self.condensed_species_indices)

    def formula_matrix_gas(self) -> Array:
        """Formula matrix for gas species

        Only columns in the formula matrix referring to gas species have non-zero values.

        Returns:
            Formula matrix for gas species
        """
        return self._get_modified_formula_matrix(self.gas_species_indices)

    def _get_number_density_output(
        self, number_density: Array, molar_mass_expanded: Array, prefix: str = ""
    ) -> dict[str, Array]:
        """Gets the outputs associated with a number density

        Args:
            number_density: Number density
            molar_mass_expanded: Molar mass associated with the number density
            prefix: Key prefix for the output. Defaults to an empty string.

        Returns
            Dictionary of output quantities
        """
        atmosphere_volume: Array = self.atmosphere_volume()
        # Volume must be a column vector because it multiples all elements in the row
        number: Array = number_density * atmosphere_volume[:, jnp.newaxis]
        moles: Array = number / AVOGADRO
        mass: Array = moles * molar_mass_expanded

        out: dict[str, Array] = {}
        out[f"{prefix}number_density"] = number_density
        out[f"{prefix}number"] = number
        out[f"{prefix}moles"] = moles
        out[f"{prefix}mass"] = mass

        return out

    def gas_species_asdict(self) -> dict[str, dict[str, Array]]:
        """Gets the gas species output as a dictionary

        Returns:
            Gas species output as a dictionary
        """
        molar_mass: Array = self.species_molar_mass_expanded()
        number_density_gas: Array = jnp.take(
            self.number_density(), self.gas_species_indices, axis=1
        )
        dissolved_gas: Array = jnp.take(
            self.species_density_in_melt(), self.gas_species_indices, axis=1
        )
        total_gas: Array = number_density_gas + dissolved_gas
        activity_gas: Array = jnp.take(self.activity(), self.gas_species_indices, axis=1)
        pressure_gas: Array = jnp.take(self.pressure(), self.gas_species_indices, axis=1)
        molar_mass_gas: Array = jnp.take(molar_mass, self.gas_species_indices, axis=1)
        gas_species: tuple[Species, ...] = tuple(
            self._species[ii] for ii in self.gas_species_indices
        )

        out: dict[str, Array] = {}
        out |= self._get_number_density_output(number_density_gas, molar_mass_gas, "atmosphere_")
        out |= self._get_number_density_output(dissolved_gas, molar_mass_gas, "dissolved_")
        out |= self._get_number_density_output(total_gas, molar_mass_gas, "total_")
        out["molar_mass"] = molar_mass
        out["volume_mixing_ratio"] = out["atmosphere_number"] / jnp.sum(
            out["atmosphere_number"], axis=1, keepdims=True
        )
        out["atmosphere_ppm"] = out["volume_mixing_ratio"] * mega
        out["atmosphere_ppmw"] = (
            out["atmosphere_mass"] / jnp.sum(out["atmosphere_mass"], axis=1, keepdims=True) * mega
        )
        out["pressure"] = pressure_gas
        out["fugacity"] = activity_gas
        out["fugacity_coefficient"] = activity_gas / pressure_gas
        out["dissolved_ppmw"] = self.species_ppmw_in_melt()

        split_dict: list[dict[str, Array]] = split_dict_by_columns(out)
        species_out: dict[str, dict[str, Array]] = {
            species.name: split_dict[ii] for ii, species in enumerate(gas_species)
        }

        return species_out

    def log_activity(self) -> Array:
        """Gets log activity of all species.

        This is usually what the user wants when referring to activity because it includes a
        consideration of species stability

        Returns:
            Log activity of all species
        """
        log_activity_without_stability: Array = self.log_activity_without_stability()
        # We select the relevant log_activity at the end of this method depending which species
        # also have stability, so its OK to build a padded array here to keep the size the same.
        log_stability_padded: Array = jnp.zeros_like(log_activity_without_stability)
        log_stability_padded = log_stability_padded.at[
            :, self._species.get_stability_species_indices()
        ].set(self.log_stability)
        # Note that the below array contains incorrect entries for species without stability since
        # exp(0)=1. But these are filtered out further below.
        log_activity_with_stability: Array = log_activity_without_stability - jnp.exp(
            log_stability_padded
        )

        # Now select the appropriate activity for each species, depending if stability is relevant.
        condition_broadcasted = jnp.broadcast_to(
            self.stability_species_mask, log_activity_without_stability.shape
        )
        # logger.debug("condition_broadcasted = %s", condition_broadcasted)

        log_activity: Array = jnp.where(
            condition_broadcasted,
            log_activity_with_stability,
            log_activity_without_stability,
        )

        return log_activity

    def log_activity_without_stability(self) -> Array:
        """Gets log activity without stability of all species

        Args:
            Log activity without stability of all species
        """
        log_activity_func: Callable = eqx.filter_vmap(
            get_log_activity,
            in_axes=(vmap_axes_spec(self._traced_parameters), None, 0),
        )
        log_activity: Array = log_activity_func(
            self._traced_parameters, self._fixed_parameters, self.log_number_density
        )

        return log_activity

    def number_density(self) -> Array:
        r"""Gets number density of all species

        Returns:
            Number density in :math:`\mathrm{molecules}\, \mathrm{m}^{-3}`
        """
        return jnp.exp(self.log_number_density)

    def species_molar_mass_expanded(self) -> Array:
        r"""Gets molar mass of all species in an expanded array.

        Returns:
            Molar mass of all species in an expanded array.
        """
        return jnp.tile(self.molar_mass, (self.number_solutions, 1))

    def pressure(self) -> Array:
        """Gets pressure of species in bar

        This will compute pressure of all species, including condensates, for simplicity.

        Returns:
            Pressure of species in bar
        """
        pressure_func: Callable = eqx.filter_vmap(
            get_pressure_from_log_number_density, in_axes=(0, self.temperature_vmap)
        )
        pressure: Array = pressure_func(self.log_number_density, self.temperature)

        return pressure

    def quick_look(self) -> dict[str, ArrayLike]:
        """Quick look at the solution

        Provides a quick first glance at the output with convenient units and to ease comparison
        with test or benchmark data.

        Returns:
            Dictionary of the solution
        """
        out: dict[str, ArrayLike] = {}

        for nn, species_ in enumerate(self._species):
            pressure: Array = self.pressure()[:, nn]
            activity: Array = self.activity()[:, nn]
            out[species_.name] = pressure
            out[f"{species_.name}_activity"] = activity

        return collapse_single_entry_values(out)

    def raw_solution_asdict(self) -> dict[str, Array]:
        """Gets the raw solution

        Returns:
            Dictionary of the raw solution
        """
        raw_solution: dict[str, Array] = {}

        species_names: tuple[str, ...] = self._species.get_species_names()

        for ii, species_name in enumerate(species_names):
            raw_solution[species_name] = self.log_number_density[:, ii]

        for ii, index in enumerate(self._species.get_stability_species_indices()):
            species_name = species_names[index]
            raw_solution[f"{species_name}_stability"] = self.log_stability[:, ii]

        return raw_solution

    def residual_asdict(self) -> dict[int, Array]:
        """Gets the residual

        Returns:
            Dictionary of the residual
        """
        residual_func: Callable = eqx.filter_vmap(
            objective_function,
            in_axes=(
                0,
                {
                    "traced_parameters": self.traced_parameters_vmap,
                    "fixed_parameters": None,
                },
            ),
        )
        residual: Array = residual_func(
            self._solution,
            {
                "traced_parameters": self._traced_parameters,
                "fixed_parameters": self._fixed_parameters,
            },
        )

        out: dict[int, Array] = {}
        for ii in range(residual.shape[1]):
            out[ii] = residual[:, ii]

        return out

    def species_density_in_melt(self) -> Array:
        """Gets species number density in the melt

        Returns:
            Species number density in the melt
        """
        species_density_in_melt_func: Callable = eqx.filter_vmap(
            get_species_density_in_melt,
            in_axes=(self.traced_parameters_vmap, None, 0, 0, 0),
        )
        species_density_in_melt: Array = species_density_in_melt_func(
            self._traced_parameters,
            self._fixed_parameters,
            self.log_number_density,
            self.log_activity(),
            self.atmosphere_log_volume(),
        )

        return species_density_in_melt

    def species_ppmw_in_melt(self) -> Array:
        """Gets species ppmw in the melt

        Return:
            Species ppmw in the melt
        """
        species_ppmw_in_melt_func: Callable = eqx.filter_vmap(
            get_species_ppmw_in_melt, in_axes=(self.traced_parameters_vmap, None, 0, 0)
        )
        species_ppmw_in_melt: Array = species_ppmw_in_melt_func(
            self._traced_parameters,
            self._fixed_parameters,
            self.log_number_density,
            self.log_activity(),
        )

        return species_ppmw_in_melt

    def stability(self) -> Array:
        """Gets stability of relevant species

        Returns:
            Stability of relevant species
        """
        return jnp.exp(self.log_stability)

    def _drop_unsuccessful_solves(
        self, dataframes: dict[str, pd.DataFrame]
    ) -> dict[str, pd.DataFrame]:
        """Drops unsuccessful solves

        Args:
            dataframes: Dataframes from which to drop unsuccessful models

        Returns:
            Dictionary of dataframes without unsuccessful models
        """
        return {key: df.loc[np.array(self._solver_status)] for key, df in dataframes.items()}

    def to_dataframes(self, drop_unsuccessful: bool = False) -> dict[str, pd.DataFrame]:
        """Gets the output in a dictionary of dataframes.

        Args:
            drop_unsuccessful: Drop models that did not solve. Defaults to False.

        Returns:
            Output in a dictionary of dataframes
        """
        if self._cached_dataframes is not None:
            logger.debug("Returning cached to_dataframes output")
            dataframes: dict[str, pd.DataFrame] = self._cached_dataframes  # Return cached result
        else:
            logger.info("Computing to_dataframes output")
            dataframes = nested_dict_to_dataframes(self.asdict())
            self._cached_dataframes = dataframes
            logger.debug("to_dataframes = %s", self._cached_dataframes)

        if drop_unsuccessful:
            logger.info("Dropping models that did not solve")
            dataframes: dict[str, pd.DataFrame] = self._drop_unsuccessful_solves(dataframes)

        return dataframes

    def to_excel(
        self, file_prefix: Path | str = "new_atmodeller_out", drop_unsuccessful: bool = False
    ) -> None:
        """Writes the output to an Excel file.

        Args:
            file_prefix: Prefix of the output file. Defaults to new_atmodeller_out.
            drop_unsuccessful: Drop models that did not solve. Defaults to False.
        """
        logger.info("Writing output to excel")
        out: dict[str, pd.DataFrame] = self.to_dataframes(drop_unsuccessful)
        output_file: Path = Path(f"{file_prefix}.xlsx")

        # Convenient to highlight rows where the solver failed to find a solution for follow-up
        # analysis
        # Define a fill color for highlighting rows (e.g., yellow)
        highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # Get the indices where the successful_solves mask is False
        unsuccessful_indices: npt.NDArray[np.int_] = np.where(
            np.array(self._solver_status) == False  # noqa: E712
        )[0]

        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            for df_name, df in out.items():
                df.to_excel(writer, sheet_name=df_name, index=True)
                sheet = writer.sheets[df_name]

                # Apply highlighting to the rows where the solver failed to find a solution
                for idx in unsuccessful_indices:
                    # Highlight the entire row (starting from index 2 to skip header row)
                    for col in range(1, len(df.columns) + 2):
                        # row=idx+2 because Excel is 1-indexed and row 1 is the header
                        cell = sheet.cell(row=idx + 2, column=col)
                        cell.fill = highlight_fill

        logger.info("Output written to %s", output_file)

    def to_pickle(
        self, file_prefix: Path | str = "new_atmodeller_out", drop_unsuccessful: bool = False
    ) -> None:
        """Writes the output to a pickle file.

        Args:
            file_prefix: Prefix of the output file. Defaults to new_atmodeller_out.
            drop_unsuccessful: Drop models that did not solve. Defaults to False.
        """
        logger.info("Writing output to pickle")
        out: dict[str, pd.DataFrame] = self.to_dataframes(drop_unsuccessful)
        output_file: Path = Path(f"{file_prefix}.pkl")

        with open(output_file, "wb") as handle:
            pickle.dump(out, handle, protocol=pickle.HIGHEST_PROTOCOL)

        logger.info("Output written to %s", output_file)


def collapse_single_entry_values(
    input_dict: dict[str, ArrayLike],
) -> dict[str, ArrayLike]:
    """Collapses single entry values in a dictionary

    Args:
        input_dict: Input dictionary

    Returns:
        Dictionary with collapsed values
    """
    out: dict[str, ArrayLike] = {}
    for key, value in input_dict.items():
        try:
            if value.size > 1:  # type: ignore because AttributeError dealt with
                out[key] = jnp.squeeze(value)
            else:
                out[key] = value.item()  # type:ignore because AttributeError dealt with
        except AttributeError:
            out[key] = value

    return out


def expand_dict(some_dict: dict[str, ArrayLike], expand_to_size: int) -> dict[str, Array]:
    """Gets a dictionary of the values with scalars expanded

    Args:
        some_dict: Some dictionary
        expand_to_size: Size to expand the arrays to

    Returns:
        A dictionary with values expanded to `expand_to_size`
    """

    def expand_to_match_size(x: ArrayLike, size: int) -> ArrayLike:
        """Expands an array

        Args:
            x: Value to possibly expand
            size: Size to expand to

        Returns:
            Expanded value
        """
        if jnp.isscalar(x):
            return jnp.broadcast_to(x, size)

        return x

    expanded_dict: dict[str, Array] = tree_map(
        lambda x: jnp.asarray(expand_to_match_size(x, expand_to_size)), some_dict
    )

    return expanded_dict


def split_dict_by_columns(dict_to_split: dict[str, Array]) -> list[dict[str, Array]]:
    """Splits a dictionary based on columns in the values.

    Args:
        dict_to_split: A dictionary to split

    Returns:
        A list of dictionaries split by column
    """
    # Get the number of columns from the first array in the dictionary
    first_key: str = next(iter(dict_to_split))
    logger.debug("first_key = %s", first_key)
    num_columns: int = dict_to_split[first_key].shape[1]

    # Function to split an array into a list of its columns
    def split_columns(array: Array) -> list[Array]:
        return [array[:, i] for i in range(num_columns)]

    # Apply the splitting function to each array in the dictionary
    split_values: dict[str, list[Array]] = jax.tree_util.tree_map(split_columns, dict_to_split)

    # Initialize a list of dictionaries, one for each column
    split_dicts: list[dict] = [{} for _ in range(num_columns)]

    # Fill the dictionaries with the corresponding columns
    for key, columns in split_values.items():
        for i, column in enumerate(columns):
            split_dicts[i][key] = column

    return split_dicts


def nested_dict_to_dataframes(
    nested_dict: dict[str, dict[str, Any]],
) -> dict[str, pd.DataFrame]:
    """Creates a dictionary of dataframes from a nested dictionary

    Args:
        nested_dict: A nested dictionary

    Returns:
        A dictionary of dataframes
    """
    dataframes: dict[str, pd.DataFrame] = {}

    for outer_key, inner_dict in nested_dict.items():
        # Convert inner dictionary to DataFrame
        df: pd.DataFrame = pd.DataFrame(inner_dict)
        dataframes[outer_key] = df

    return dataframes
