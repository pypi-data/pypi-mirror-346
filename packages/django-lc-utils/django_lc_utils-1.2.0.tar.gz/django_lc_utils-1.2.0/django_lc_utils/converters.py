import json
import re
from io import BytesIO

from html_sanitizer import Sanitizer
from html_sanitizer.django import get_sanitizer
from openpyxl import load_workbook

from .sanitizers import HTML_SANITIZER_CUSTOM_SETTINGS


def xlsx_to_json(xlsx_file, group_by_key=None):
    wb = load_workbook(BytesIO(xlsx_file.read()))
    # Iterate through each row in worksheet and fetch values into dict
    sheet_obj = wb.active

    # Cell objects also have a row, column,
    # and coordinate attributes that provide
    # location information for the cell.

    # Note: The first row or
    # column integer is 1, not 0.
    # cell_obj = sheet_obj.cell(row=1, column=1)

    group_key_idx = None
    HEADERS = []
    for col_idx in range(1, sheet_obj.max_column + 1):
        data = sheet_obj.cell(row=1, column=col_idx).value
        HEADERS.append(data)
        if group_by_key == data:
            group_key_idx = col_idx

    if group_key_idx:
        data_list = {}
    else:
        data_list = []

    for row_idx in range(2, sheet_obj.max_row + 1):
        row_obj = {}
        key = sheet_obj.cell(row=row_idx, column=group_key_idx).value if group_key_idx else None

        for header_idx, col_idx in enumerate(range(1, sheet_obj.max_column + 1)):
            if col_idx != group_key_idx:
                row_obj[HEADERS[header_idx]] = sheet_obj.cell(row=row_idx, column=col_idx).value

        if group_key_idx:
            if key not in data_list:
                data_list[key] = []
            data_list[key].append(row_obj)
        else:
            data_list.append(row_obj)

    # Serialize the list of dicts to JSON
    return json.dumps(data_list)


# TODO: Need the path as args for this so that it can be reused accross projects
# def html_to_pdf(html, stylesheets=None):
#     bio = BytesIO()
#     ROOT_DIR = Path(__file__).resolve(strict=True).parent.parent.parent
#     APPS_DIR = str(ROOT_DIR / "los/templates")
#     if stylesheets is None:
#         stylesheets = [CSS(APPS_DIR + "/css/los_app.css")]
#     else:
#         stylesheets = [CSS(APPS_DIR + f"/css/{stylesheets}")]

#     kwargs = {"string": html}
#     # if stylesheets:
#     #     kwargs["stylesheets"] = stylesheets

#     HTML(**kwargs).write_pdf(bio, stylesheets=stylesheets)
#     bio.seek(0)
#     return bio


# NOTE: imported a library inside a function, why so?
def convert_numbers_to_words(number):
    import inflect

    p = inflect.engine()

    number_words = p.number_to_words(number)
    return number_words


def parse_json_string(value):
    try:
        value = json.loads(value)
    except json.decoder.JSONDecodeError:
        value
    return value if value != "-" else None


def snake_case(s: str) -> str:
    """Utility function to convert strings in CamelCase to snake_case

    Args:
        s (str): input string

    Returns:
        str: snake_case string
    """
    s = re.sub("[^0-9a-zA-Z]+", "", s)
    return "_".join(re.sub("([A-Z][a-z]+)", r" \1", re.sub("([A-Z]+)", r" \1", s.replace("-", " "))).split()).lower()


def phone_format(number: str) -> str:
    """Format phone number with dashes

    Args:
        number (str): unformatted phone number

    Returns:
        str: formatted phone number ie 555-555-5555
    """
    number = re.sub("[^0-9]", "", number)
    number = number[-10:]
    number = f"{number[:3]}-{number[3:6]}-{number[6:]}"
    return number


def clean_html(value: str, sanitizer="default") -> str:
    """Helper function to sanitize HTML to prevent cross-site scripting vulnerabilities.  If no `sanitizer`
    is specified, we'll default to the "default" sanitizer, and pass in our own custom settings from
    `settings.HTML_SANITIZER_CUSTOM_SETTINGS`

    Args:
        value (str): input html string
        sanitizer (str, optional): Sanitizer name to pass to `get_sanitizer` function. Defaults to "default"

    Returns:
        str: cleaned html string
    """
    if not isinstance(value, str):
        return value
    if sanitizer == "default":
        # use the default serializer with custom settings
        sanitize_obj = Sanitizer(HTML_SANITIZER_CUSTOM_SETTINGS)
    else:
        # use the specified serializer
        sanitize_obj = get_sanitizer(name=sanitizer)

    return sanitize_obj.sanitize(value)
